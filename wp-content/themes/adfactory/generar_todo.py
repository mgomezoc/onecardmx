#!/usr/bin/env python3
"""
Script completo para procesar Gasolina.xlsx
- Genera estadosData.js para el mapa
- Genera archivos CSV/Excel individuales por estado para descargar
- Actualiza automáticamente todo desde un solo archivo maestro

Uso: python3 generar_todo.py
"""

import openpyxl
from openpyxl import Workbook
import json
import csv
import os
from collections import OrderedDict

# Configuración
FORMATO_SALIDA = 'csv'  # 'csv' o 'xlsx' - CSV es más rápido y ligero
URL_BASE_DESCARGAS = 'http://onecard.mx/wp-content/themes/adfactory/descargas_estados/'

# Coordenadas de cada estado (centro geográfico)
COORDENADAS = {
    'AGUASCALIENTES': [-102.2916, 21.8853],
    'BAJA CALIFORNIA NORTE': [-116.6234, 30.8406],
    'BAJA CALIFORNIA SUR': [-111.6671, 26.0444],
    'CAMPECHE': [-90.5349, 19.8301],
    'CHIAPAS': [-92.6376, 16.7569],
    'CHIHUAHUA': [-106.0691, 28.6330],
    'CIUDAD DE MEXICO': [-99.1332, 19.4326],
    'COAHUILA': [-101.7068, 27.0587],
    'COLIMA': [-103.7240, 19.2452],
    'DURANGO': [-104.6532, 24.0277],
    'GUANAJUATO': [-101.2574, 21.0190],
    'GUERRERO': [-99.5451, 17.4392],
    'HIDALGO': [-98.7624, 20.0911],
    'JALISCO': [-103.3494, 20.6595],
    'MEXICO': [-99.6713, 19.2832],
    'MICHOACAN': [-101.7068, 19.5666],
    'MORELOS': [-99.2233, 18.6813],
    'NAYARIT': [-104.8936, 21.7514],
    'NUEVO LEON': [-100.3161, 25.5922],
    'OAXACA': [-96.7266, 17.0732],
    'PUEBLA': [-98.2063, 19.0414],
    'QUERETARO': [-100.3899, 20.5888],
    'QUINTANA ROO': [-88.2963, 19.1817],
    'SAN LUIS POTOSI': [-100.9855, 22.1565],
    'SINALOA': [-107.3940, 25.8243],
    'SONORA': [-110.9559, 29.2972],
    'TABASCO': [-92.9475, 17.8409],
    'TAMAULIPAS': [-99.1013, 24.2669],
    'TLAXCALA': [-98.2375, 19.3139],
    'VERACRUZ': [-96.1342, 19.1738],
    'YUCATAN': [-89.5926, 20.7099],
    'ZACATECAS': [-102.5832, 22.7709]
}

# Mapeo de nombres de estados a keys
ESTADO_KEYS = {
    'AGUASCALIENTES': 'aguascalientes',
    'BAJA CALIFORNIA NORTE': 'baja-california',
    'BAJA CALIFORNIA SUR': 'baja-california-sur',
    'CAMPECHE': 'campeche',
    'CHIAPAS': 'chiapas',
    'CHIHUAHUA': 'chihuahua',
    'CIUDAD DE MEXICO': 'cdmx',
    'COAHUILA': 'coahuila',
    'COLIMA': 'colima',
    'DURANGO': 'durango',
    'GUANAJUATO': 'guanajuato',
    'GUERRERO': 'guerrero',
    'HIDALGO': 'hidalgo',
    'JALISCO': 'jalisco',
    'MEXICO': 'mexico',
    'MICHOACAN': 'michoacan',
    'MORELOS': 'morelos',
    'NAYARIT': 'nayarit',
    'NUEVO LEON': 'nuevo-leon',
    'OAXACA': 'oaxaca',
    'PUEBLA': 'puebla',
    'QUERETARO': 'queretaro',
    'QUINTANA ROO': 'quintana-roo',
    'SAN LUIS POTOSI': 'san-luis-potosi',
    'SINALOA': 'sinaloa',
    'SONORA': 'sonora',
    'TABASCO': 'tabasco',
    'TAMAULIPAS': 'tamaulipas',
    'TLAXCALA': 'tlaxcala',
    'VERACRUZ': 'veracruz',
    'YUCATAN': 'yucatan',
    'ZACATECAS': 'zacatecas'
}

# Mapeo de keys a nombres para display
NOMBRES_DISPLAY = {
    'aguascalientes': 'Aguascalientes',
    'baja-california': 'Baja California',
    'baja-california-sur': 'Baja California Sur',
    'campeche': 'Campeche',
    'chiapas': 'Chiapas',
    'chihuahua': 'Chihuahua',
    'cdmx': 'Ciudad de México',
    'coahuila': 'Coahuila',
    'colima': 'Colima',
    'durango': 'Durango',
    'guanajuato': 'Guanajuato',
    'guerrero': 'Guerrero',
    'hidalgo': 'Hidalgo',
    'jalisco': 'Jalisco',
    'mexico': 'Estado de México',
    'michoacan': 'Michoacán',
    'morelos': 'Morelos',
    'nayarit': 'Nayarit',
    'nuevo-leon': 'Nuevo León',
    'oaxaca': 'Oaxaca',
    'puebla': 'Puebla',
    'queretaro': 'Querétaro',
    'quintana-roo': 'Quintana Roo',
    'san-luis-potosi': 'San Luis Potosí',
    'sinaloa': 'Sinaloa',
    'sonora': 'Sonora',
    'tabasco': 'Tabasco',
    'tamaulipas': 'Tamaulipas',
    'tlaxcala': 'Tlaxcala',
    'veracruz': 'Veracruz',
    'yucatan': 'Yucatán',
    'zacatecas': 'Zacatecas'
}

# Mapeo de keys a nombres de archivos
ARCHIVOS_DESCARGA = {
    'aguascalientes': 'aguascalientes',
    'baja-california': 'baja_california_norte',
    'baja-california-sur': 'baja_california_sur',
    'campeche': 'campeche',
    'chiapas': 'chiapas',
    'chihuahua': 'chihuahua',
    'cdmx': 'ciudad_de_mexico',
    'coahuila': 'coahuila',
    'colima': 'colima',
    'durango': 'durango',
    'guanajuato': 'guanajuato',
    'guerrero': 'guerrero',
    'hidalgo': 'hidalgo',
    'jalisco': 'jalisco',
    'mexico': 'mexico',
    'michoacan': 'michoacan',
    'morelos': 'morelos',
    'nayarit': 'nayarit',
    'nuevo-leon': 'nuevo_leon',
    'oaxaca': 'oaxaca',
    'puebla': 'puebla',
    'queretaro': 'queretaro',
    'quintana-roo': 'quintana_roo',
    'san-luis-potosi': 'san_luis_potosi',
    'sinaloa': 'sinaloa',
    'sonora': 'sonora',
    'tabasco': 'tabasco',
    'tamaulipas': 'tamaulipas',
    'tlaxcala': 'tlaxcala',
    'veracruz': 'veracruz',
    'yucatan': 'yucatan',
    'zacatecas': 'zacatecas'
}

def procesar_excel(archivo='Gasolina.xlsx'):
    """Procesa el archivo Excel y agrupa comercios por estado"""
    print(f"📂 Leyendo {archivo}...")
    
    # Cargar Excel
    wb = openpyxl.load_workbook(archivo, read_only=True)
    ws = wb.active
    
    # Obtener headers
    headers = [cell.value for cell in ws[1]]
    
    # Agrupar comercios por estado
    comercios_por_estado = {}
    total_comercios = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        estado = row[7]  # Columna "Estado"
        
        if estado:
            if estado not in comercios_por_estado:
                comercios_por_estado[estado] = []
            
            comercios_por_estado[estado].append(row)
            total_comercios += 1
    
    wb.close()
    
    print(f"✓ Procesados {total_comercios:,} comercios en {len(comercios_por_estado)} estados")
    
    return headers, comercios_por_estado

def generar_archivos_por_estado(headers, comercios_por_estado, formato='csv'):
    """Genera archivos individuales por estado (CSV o Excel)"""
    
    # Crear directorio de salida
    output_dir = 'descargas_estados'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    print(f"\n📦 Generando archivos {formato.upper()} por estado...")
    print("=" * 60)
    
    archivos_generados = {}
    
    for estado_nombre, comercios in sorted(comercios_por_estado.items()):
        key = ESTADO_KEYS.get(estado_nombre)
        if not key:
            print(f"⚠ Advertencia: Estado '{estado_nombre}' no tiene key definida")
            continue
        
        nombre_archivo = ARCHIVOS_DESCARGA[key]
        
        if formato == 'csv':
            # Generar CSV
            archivo_path = os.path.join(output_dir, f'{nombre_archivo}.csv')
            
            with open(archivo_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(headers)  # Escribir headers
                writer.writerows(comercios)  # Escribir datos
            
        else:  # xlsx
            # Generar Excel
            archivo_path = os.path.join(output_dir, f'{nombre_archivo}.xlsx')
            
            wb = Workbook()
            ws = wb.active
            ws.title = estado_nombre[:31]  # Límite de Excel
            
            # Escribir headers
            ws.append(headers)
            
            # Escribir datos
            for row in comercios:
                ws.append(row)
            
            wb.save(archivo_path)
        
        archivos_generados[key] = {
            'archivo': f'{nombre_archivo}.{formato}',
            'path': archivo_path,
            'cantidad': len(comercios)
        }
        
        print(f"✓ {estado_nombre}: {len(comercios):,} comercios → {nombre_archivo}.{formato}")
    
    print("=" * 60)
    print(f"✓ {len(archivos_generados)} archivos generados en '{output_dir}/'")
    
    return archivos_generados

def generar_javascript(comercios_por_estado, archivos_generados, formato='csv'):
    """Genera el código JavaScript para estadosData"""
    
    output = []
    output.append("const estadosData = {")
    
    for estado_nombre in sorted(comercios_por_estado.keys()):
        key = ESTADO_KEYS.get(estado_nombre)
        if not key:
            continue
        
        nombre_display = NOMBRES_DISPLAY[key]
        coords = COORDENADAS.get(estado_nombre, [-99.1332, 19.4326])
        cantidad = len(comercios_por_estado[estado_nombre])
        archivo = archivos_generados[key]['archivo']
        
        output.append(f"  '{key}': {{")
        output.append(f"    nombre: '{nombre_display}',")
        output.append(f"    estaciones: '{cantidad:,}+',")
        output.append(f"    descripcion: 'Red completa de estaciones en el estado',")
        output.append(f"    url: '{URL_BASE_DESCARGAS}{archivo}',")
        output.append(f"    coordinates: [{coords[0]}, {coords[1]}]")
        output.append(f"  }},")
    
    output.append("};")
    
    return "\n".join(output)

def generar_json(comercios_por_estado, archivos_generados, formato='csv'):
    """Genera archivo JSON con los datos"""
    
    data = OrderedDict()
    
    for estado_nombre in sorted(comercios_por_estado.keys()):
        key = ESTADO_KEYS.get(estado_nombre)
        if not key:
            continue
        
        nombre_display = NOMBRES_DISPLAY[key]
        coords = COORDENADAS.get(estado_nombre, [-99.1332, 19.4326])
        cantidad = len(comercios_por_estado[estado_nombre])
        archivo = archivos_generados[key]['archivo']
        
        data[key] = {
            'nombre': nombre_display,
            'estaciones': f'{cantidad:,}+',
            'descripcion': 'Red completa de estaciones en el estado',
            'url': f'{URL_BASE_DESCARGAS}{archivo}',
            'coordinates': coords
        }
    
    return json.dumps(data, indent=2, ensure_ascii=False)

def main():
    print("=" * 60)
    print("🚀 GENERADOR COMPLETO DE ARCHIVOS PARA MAPA")
    print("=" * 60)
    print()
    
    # Procesar Excel maestro
    headers, comercios_por_estado = procesar_excel('Gasolina.xlsx')
    
    # Generar archivos por estado
    archivos_generados = generar_archivos_por_estado(
        headers, 
        comercios_por_estado, 
        formato=FORMATO_SALIDA
    )
    
    print()
    print("=" * 60)
    print("📝 Generando código JavaScript...")
    print("=" * 60)
    
    # Generar JavaScript
    js_code = generar_javascript(comercios_por_estado, archivos_generados, FORMATO_SALIDA)
    
    with open('estadosData.js', 'w', encoding='utf-8') as f:
        f.write(js_code)
    
    print("✓ Archivo generado: estadosData.js")
    
    # Generar JSON
    print()
    print("=" * 60)
    print("📝 Generando archivo JSON...")
    print("=" * 60)
    
    json_data = generar_json(comercios_por_estado, archivos_generados, FORMATO_SALIDA)
    
    with open('comercios-combustible.json', 'w', encoding='utf-8') as f:
        f.write(json_data)
    
    print("✓ Archivo generado: comercios-combustible.json")
    
    # Mostrar resumen
    print()
    print("=" * 60)
    print("📊 RESUMEN:")
    print("=" * 60)
    total_comercios = sum(len(v) for v in comercios_por_estado.values())
    print(f"Total de comercios: {total_comercios:,}")
    print(f"Total de estados: {len(comercios_por_estado)}")
    print(f"Formato de archivos: {FORMATO_SALIDA.upper()}")
    print()
    
    print("Top 5 estados con más comercios:")
    top_estados = sorted(
        comercios_por_estado.items(), 
        key=lambda x: len(x[1]), 
        reverse=True
    )[:5]
    
    for i, (estado, comercios) in enumerate(top_estados, 1):
        print(f"  {i}. {estado}: {len(comercios):,} comercios")
    
    print()
    print("=" * 60)
    print("✅ ARCHIVOS GENERADOS:")
    print("=" * 60)
    print(f"1. descargas_estados/ - {len(archivos_generados)} archivos {FORMATO_SALIDA.upper()} por estado")
    print("2. estadosData.js - Para copiar/pegar en tu JavaScript")
    print("3. comercios-combustible.json - Para cargar dinámicamente")
    print()
    
    print("=" * 60)
    print("📤 PRÓXIMO PASO:")
    print("=" * 60)
    print(f"Sube la carpeta 'descargas_estados/' a:")
    print(f"  {URL_BASE_DESCARGAS}")
    print()
    print("Y el mapa cargará automáticamente los archivos correctos!")
    print("=" * 60)
    print()
    print("✨ ¡Proceso completado exitosamente!")

if __name__ == '__main__':
    main()
